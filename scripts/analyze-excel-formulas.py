import openpyxl
import json

# Load the Excel file
file_path = r"C:\Users\jonas\Downloads\V-0101 Matrix Vessel NII Coverage Calculator MP version2.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

# Get all sheet names
print("Sheets in workbook:", wb.sheetnames)
print("\n" + "="*50 + "\n")

# Analyze each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"SHEET: {sheet_name}")
    print("-"*30)

    # Find cells with formulas
    formulas = {}
    values = {}

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_addr = f"{cell.column_letter}{cell.row}"

                # Check if it's a formula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas[cell_addr] = cell.value
                    # Also try to get the calculated value
                    wb_data = openpyxl.load_workbook(file_path, data_only=True)
                    sheet_data = wb_data[sheet_name]
                    calculated_value = sheet_data[cell_addr].value
                    values[cell_addr] = calculated_value
                elif cell.value not in ['', None]:
                    # Store non-empty values for reference
                    values[cell_addr] = cell.value

    # Print important input cells (usually have labels)
    print("\nKey Input/Output cells:")
    for addr, val in values.items():
        if isinstance(val, str) and any(keyword in val.lower() for keyword in
            ['tan', 'diameter', 'id', 'wall', 'wt', 'circ', 'area', 'grid', 'hour',
             'segment', 'clock', 'shell', 'dome', 'paut', 'pec', 'tofd', 'length',
             'axis', 'access', 'scan', 'analysis', 'report', 'total']):
            print(f"  {addr}: {val}")

    print("\nKey Formulas:")
    # Print formulas with their addresses
    for addr, formula in formulas.items():
        print(f"  {addr}: {formula}")
        if addr in values:
            print(f"       -> Value: {values[addr]}")

    print("\n" + "="*50 + "\n")

# Close the workbook
wb.close()